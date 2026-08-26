"""inspect 端點的整合測試（批次匯入 UI 規劃路線 A）。

只做「讀檔 + 偵測欄位」，不轉成 rows、不寫入——這裡驗證的是端點層的
組裝與 HTTP 行為；_detect_columns()/read_grid() 純邏輯已在
test_alarm_ingest.py 涵蓋。
"""
import io

import openpyxl
import pytest


def _xlsx_file(rows: list, filename: str = "vendor.xlsx", sheet_name: str = "Sheet1"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return (buf, filename)


def _xlsx_multi_sheet(sheets: dict, filename: str = "vendor.xlsx"):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(name)
        for row in rows:
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return (buf, filename)


VENDOR_ROWS = [
    ["Description", "Cause", "Action"],
    ["0024 Forming Panel Heating Units Automatic", "circuit breaker tripped", "Check sensors"],
    ["0025 Sealing Panel Heating Units Automatic", "circuit breaker tripped", "Check sensors"],
    ["0033 Area 1 Running", "unit is running", ""],
    ["0035 Manual Stop", "stop button pressed", "Check button"],
    ["0044 No Compressed Air", "insufficient pressure", "Check pneumatic circuit"],
]


def test_inspect_requires_admin(anon_client):
    r = anon_client.post(
        "/api/admin/import/local/inspect",
        data={"file": _xlsx_file(VENDOR_ROWS)},
        content_type="multipart/form-data",
    )
    assert r.status_code in (302, 401, 403)


def test_inspect_missing_file_rejected(client):
    r = client.post("/api/admin/import/local/inspect", data={}, content_type="multipart/form-data")
    assert r.status_code == 400


def test_inspect_unsupported_extension_rejected(client):
    r = client.post(
        "/api/admin/import/local/inspect",
        data={"file": (io.BytesIO(b"hello"), "notes.txt")},
        content_type="multipart/form-data",
    )
    assert r.status_code == 400


def test_inspect_rejects_pdf_and_docx(client):
    """PDF/Word 目前仍只在 CLI 走 tools/variant/parse_alarms.py，
    inspect 端點不開放這兩種格式上傳。"""
    for filename in ("manual.pdf", "manual.docx"):
        r = client.post(
            "/api/admin/import/local/inspect",
            data={"file": (io.BytesIO(b"fake"), filename)},
            content_type="multipart/form-data",
        )
        assert r.status_code == 400, filename


def test_inspect_does_not_require_fixed_headers(client):
    """跟固定範本路徑（bulk-import/preview）的關鍵差異：這裡沒有
    REQUIRED_HEADERS 檢查，"Description/Cause/Action" 這種原廠常見
    表頭不會被擋。"""
    r = client.post(
        "/api/admin/import/local/inspect",
        data={"file": _xlsx_file(VENDOR_ROWS)},
        content_type="multipart/form-data",
    )
    assert r.status_code == 200
    body = r.get_json()
    assert body["selected"]["detected"] is True


def test_inspect_undetected_sheet_reports_diagnostic(client):
    """偵測失敗時要告訴使用者「預期什麼、實際看到什麼」，不能只回
    detected: false——管理員不會打開程式碼看 HEAD 關鍵字表是什麼。"""
    r = client.post(
        "/api/admin/import/local/inspect",
        data={"file": _xlsx_file([["Question"], ["why does it jam"], ["what to check"]])},
        content_type="multipart/form-data",
    )
    body = r.get_json()
    assert body["selected"]["detected"] is False
    assert "diagnostic" in body["selected"]
    assert "Question" in body["selected"]["diagnostic"]


def test_inspect_total_row_count_capped(client):
    """inspect 階段還沒有 rows，用 grid 列數把關（跟固定範本路徑口徑
    不同但目的一致：避免過大檔案拖垮偵測與 samples 抓取）。"""
    import app as app_module
    rows = [["Description", "Cause", "Action"]]
    for i in range(app_module.BULK_IMPORT_MAX_ROWS + 10):
        rows.append([f"{1000+i} Alarm", "cause", "action"])
    r = client.post(
        "/api/admin/import/local/inspect",
        data={"file": _xlsx_file(rows)},
        content_type="multipart/form-data",
    )
    assert r.status_code == 400


def test_inspect_does_not_write(client):
    r = client.post(
        "/api/admin/import/local/inspect",
        data={"file": _xlsx_file(VENDOR_ROWS)},
        content_type="multipart/form-data",
    )
    assert r.status_code == 200
    listed = client.get("/api/alarms").get_json()
    assert not any(a["code"] == "0024" for a in listed)


def test_inspect_reports_column_roles_and_samples(client):
    r = client.post(
        "/api/admin/import/local/inspect",
        data={"file": _xlsx_file(VENDOR_ROWS)},
        content_type="multipart/form-data",
    )
    body = r.get_json()
    cols = body["selected"]["columns"]
    by_role = {c["suggested"]: c for c in cols if c["suggested"]}

    assert by_role["code+variant"]["samples"][0].startswith("0024")
    assert "circuit breaker" in by_role["cause"]["samples"][0]
    assert "Check sensors" in by_role["action"]["samples"][0]
    # code 沒有真的被切分——inspect 只偵測欄位，不轉換內容，samples 應
    # 該是原始儲存格文字，不是切過的 code。
    assert by_role["code+variant"]["samples"][0] == "0024 Forming Panel Heating Units Automatic"


def test_inspect_sample_count_capped(client):
    rows = [["Description", "Cause", "Action"]]
    for i in range(10):
        rows.append([f"{1000+i} Some Alarm", "cause text", "action text"])
    r = client.post(
        "/api/admin/import/local/inspect",
        data={"file": _xlsx_file(rows)},
        content_type="multipart/form-data",
    )
    body = r.get_json()
    cols = body["selected"]["columns"]
    desc_col = next(c for c in cols if c["suggested"] == "code+variant")
    assert len(desc_col["samples"]) == 3


def test_inspect_multi_sheet_summary(client):
    """多分頁檔案（FILL203 那種）：sheets 清單要涵蓋全部分頁，每個
    分頁各自標記是否偵測到欄位（不是只回第一個能用的分頁）。"""
    r = client.post(
        "/api/admin/import/local/inspect",
        data={"file": _xlsx_multi_sheet({
            "alarm list": VENDOR_ROWS,
            "Problem": [["Question"], ["why does it jam"], ["what to check"]],
        })},
        content_type="multipart/form-data",
    )
    body = r.get_json()
    names = {s["name"]: s["detected"] for s in body["sheets"]}
    assert names == {"alarm list": True, "Problem": False}
    # 沒有指定 sheet 時，預設選中第一個偵測成功的分頁
    assert body["selected"]["name"] == "alarm list"


def test_inspect_select_specific_sheet_via_query_param(client):
    r = client.post(
        "/api/admin/import/local/inspect?sheet=Problem",
        data={"file": _xlsx_multi_sheet({
            "alarm list": VENDOR_ROWS,
            "Problem": [["Question"], ["why does it jam"], ["what to check"]],
        })},
        content_type="multipart/form-data",
    )
    body = r.get_json()
    assert body["selected"]["detected"] is False


def test_inspect_unknown_sheet_name_rejected(client):
    r = client.post(
        "/api/admin/import/local/inspect?sheet=NotASheet",
        data={"file": _xlsx_file(VENDOR_ROWS)},
        content_type="multipart/form-data",
    )
    assert r.status_code == 400


def test_inspect_file_too_large_rejected(client):
    """BULK_IMPORT_MAX_BYTES 是對原始上傳 bytes 長度的檢查，在讀取
    檔案內容之前就先擋（見 app.py 的 len(raw) 判斷）——用隨機亂數
    塞內容而非重複字串，因為 xlsx 的 zip 壓縮對重複內容效果太好，
    重複字串湊不出真正超過門檻的檔案大小。"""
    import random
    import string

    big_rows = [["Description", "Cause", "Action"]]
    for i in range(1, 3500):
        junk = "".join(random.choices(string.ascii_letters, k=800))
        big_rows.append([f"{i} Alarm {junk}", junk, junk])
    file_tuple = _xlsx_file(big_rows)
    assert len(file_tuple[0].getvalue()) > 2 * 1024 * 1024, "測資本身沒超過門檻，測試前提不成立"

    r = client.post(
        "/api/admin/import/local/inspect",
        data={"file": file_tuple},
        content_type="multipart/form-data",
    )
    assert r.status_code == 400


def test_inspect_rejects_nonexistent_department_via_resolve_target(client):
    """驗證的是 resolve_target_department() 這個純函式本身的行為
    （不存在的部門一律擋下），不是跨部門隔離本身——pytest 走 JsonStore
    fallback 測不到真實隔離，見 tests/test_no_fake_isolation_claims.py。
    inspect 端點呼叫這個函式只是為了複用既有的部門驗證規則。"""
    r = client.post(
        "/api/admin/import/does-not-exist-dept/inspect",
        data={"file": _xlsx_file(VENDOR_ROWS)},
        content_type="multipart/form-data",
    )
    assert r.status_code in (403, 404)
